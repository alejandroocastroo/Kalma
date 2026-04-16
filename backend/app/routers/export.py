"""Export router — Generación de reportes Excel contables"""
import io
import uuid
from datetime import date, datetime
from decimal import Decimal
from functools import lru_cache
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.auth.jwt import get_current_active_user
from app.models.payment import Payment
from app.models.client import Client
from app.models.space import Space
from app.models.tenant import Tenant

router = APIRouter(prefix="/export", tags=["Exportar"])

CATEGORY_LABELS: dict[str, str] = {
    "clase_grupal": "Clase grupal",
    "clase_privada": "Clase privada",
    "paquete_sesiones": "Paquete de sesiones",
    "membresia": "Membresía",
    "inscripcion": "Inscripción / matrícula",
    "otro_ingreso": "Otro ingreso",
    "arriendo": "Arriendo",
    "servicios_publicos": "Servicios públicos",
    "nomina_instructores": "Nómina instructores",
    "nomina_admin": "Nómina administrativa",
    "mantenimiento": "Mantenimiento",
    "equipamiento": "Equipamiento",
    "marketing": "Marketing y publicidad",
    "contabilidad": "Contabilidad",
    "tecnologia": "Software / tecnología",
    "seguros": "Seguros",
    "otros_gastos": "Otros gastos",
    # legacy keys
    "class_fee": "Cobro de clase",
    "membership": "Membresía",
    "package": "Paquete",
    "equipment": "Equipamiento",
    "rent": "Arriendo",
    "salary": "Nómina",
    "other": "Otro",
}

METHOD_LABELS: dict[str, str] = {
    "cash": "Efectivo",
    "transfer": "Transferencia",
    "card": "Tarjeta",
    "nequi": "Nequi",
    "daviplata": "Daviplata",
}

# Cached border — same for every cell, no need to instantiate per call
_THIN_BORDER = Border(
    left=Side(border_style="thin", color="D0D0D0"),
    right=Side(border_style="thin", color="D0D0D0"),
    top=Side(border_style="thin", color="D0D0D0"),
    bottom=Side(border_style="thin", color="D0D0D0"),
)


@lru_cache(maxsize=64)
def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


@lru_cache(maxsize=32)
def _font(size: int = 10, bold: bool = False, color: str = "000000") -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)


@lru_cache(maxsize=16)
def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _cop(value) -> int:
    try:
        return int(Decimal(str(value)))
    except Exception:
        return 0


def _style_header_row(ws, row: int, cols: int, bg: str = "1F4E79", fg: str = "FFFFFF"):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _font(10, bold=True, color=fg)
        cell.fill = _fill(bg)
        cell.alignment = _align("center")
        cell.border = _THIN_BORDER


def _style_total_row(ws, row: int, cols: int, bg: str = "E8EAF6"):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _font(10, bold=True)
        cell.fill = _fill(bg)
        cell.border = _THIN_BORDER


def _autofit(ws, min_width: int = 10, max_width: int = 50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(cell.value or "")) for cell in col_cells), default=0)
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


def _derive_space_columns(payments: list[Payment], space_names: dict) -> list[str]:
    """Return sorted space names present in the data, General last."""
    seen = {
        space_names.get(p.space_id, "General") if p.space_id else "General"
        for p in payments
    }
    cols = sorted(s for s in seen if s != "General")
    if "General" in seen:
        cols.append("General")
    return cols or ["General"]


def _build_resumen(ws, payments: list[Payment], tenant_name: str, start: str, end: str, space_names: dict):
    """Hoja 1: Resumen Ejecutivo — Estado de resultados del período."""
    ws.freeze_panes = "A6"

    columns = _derive_space_columns(payments, space_names)
    n_cols = len(columns) + 2  # Categoría + space cols + TOTAL
    last_col = get_column_letter(n_cols)

    # Title block
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = tenant_name
    ws["A1"].font = _font(14, bold=True, color="1F4E79")
    ws["A1"].alignment = _align("center")

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Reporte Contable — {start}  →  {end}"
    ws["A2"].font = _font(10, color="555555")
    ws["A2"].alignment = _align("center")

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = _font(9, color="888888")
    ws["A3"].alignment = _align("center")

    headers = ["Categoría"] + columns + ["TOTAL"]

    # Single pass: aggregate inc/exp by category and space
    inc_data: dict[str, dict[str, Decimal]] = {}
    exp_data: dict[str, dict[str, Decimal]] = {}
    for p in payments:
        label = CATEGORY_LABELS.get(p.category, p.category)
        space = space_names.get(p.space_id, "General") if p.space_id else "General"
        bucket = inc_data if p.type == "income" else exp_data
        if label not in bucket:
            bucket[label] = {s: Decimal(0) for s in columns}
        bucket[label][space] = bucket[label].get(space, Decimal(0)) + p.amount

    def _render_section(data, label_text, header_bg, row_fill, total_bg, total_txt_color, row):
        ws.merge_cells(f"A{row}:{last_col}{row}")
        ws[f"A{row}"] = label_text
        ws[f"A{row}"].font = _font(10, bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = _fill(header_bg)
        ws[f"A{row}"].alignment = _align("center")
        row += 1

        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, n_cols, bg=header_bg)
        row += 1

        totals = {s: Decimal(0) for s in columns}
        for cat, spaces in data.items():
            row_total = sum(spaces.get(s, Decimal(0)) for s in columns)
            values = [cat] + [_cop(spaces.get(s, 0)) for s in columns] + [_cop(row_total)]
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = _THIN_BORDER
                cell.font = _font(10)
                cell.alignment = _align("right") if c > 1 else _align("left")
                if c > 1:
                    cell.number_format = '#,##0'
                if row % 2 == 0:
                    cell.fill = _fill(row_fill)
            for s in columns:
                totals[s] += spaces.get(s, Decimal(0))
            row += 1

        grand = sum(totals.values())
        _style_total_row(ws, row, n_cols, bg=total_bg)
        ws.cell(row=row, column=1, value=f"TOTAL {label_text}").font = _font(10, bold=True)
        for c, s in enumerate(columns, 2):
            cell = ws.cell(row=row, column=c, value=_cop(totals[s]))
            cell.number_format = '#,##0'
            cell.font = _font(10, bold=True, color=total_txt_color)
            cell.alignment = _align("right")
            cell.fill = _fill(total_bg)
            cell.border = _THIN_BORDER
        cell = ws.cell(row=row, column=n_cols, value=_cop(grand))
        cell.number_format = '#,##0'
        cell.font = _font(11, bold=True, color=total_txt_color)
        cell.alignment = _align("right")
        cell.fill = _fill(total_bg)
        cell.border = _THIN_BORDER
        return row + 2, grand

    row = 5
    row, grand_inc = _render_section(inc_data, "INGRESOS", "43A047", "F1F8E9", "C8E6C9", "1B5E20", row)
    row, grand_exp = _render_section(exp_data, "EGRESOS", "E53935", "FFEBEE", "FFCDD2", "B71C1C", row)

    # Balance neto
    neto = grand_inc - grand_exp
    neto_color = "1565C0" if neto >= 0 else "B71C1C"
    neto_bg = "BBDEFB" if neto >= 0 else "FFCDD2"
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols - 1)}{row}")
    ws[f"A{row}"] = "BALANCE NETO DEL PERÍODO"
    ws[f"A{row}"].font = _font(11, bold=True)
    ws[f"A{row}"].fill = _fill("E3F2FD")
    ws[f"A{row}"].alignment = _align("right")
    ws[f"A{row}"].border = _THIN_BORDER
    cell = ws.cell(row=row, column=n_cols, value=_cop(neto))
    cell.number_format = '#,##0'
    cell.font = _font(12, bold=True, color=neto_color)
    cell.fill = _fill(neto_bg)
    cell.alignment = _align("right")
    cell.border = _THIN_BORDER

    _autofit(ws)


def _build_detalle(ws, payments: list[Payment], tipo: str, client_names: dict, space_names: dict):
    """Hoja 2/3: Detalle de ingresos o egresos."""
    ws.freeze_panes = "A2"
    filtrado = [p for p in payments if p.type == tipo]

    headers = ["Fecha", "Descripción", "Cliente", "Categoría", "Espacio", "Método de pago", "Monto (COP)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    total = Decimal(0)
    for i, p in enumerate(filtrado, 2):
        space = space_names.get(p.space_id, "General") if p.space_id else "General"
        values = [
            p.payment_date.strftime("%d/%m/%Y") if p.payment_date else "",
            p.description or "",
            client_names.get(p.client_id, "") if p.client_id else "",
            CATEGORY_LABELS.get(p.category, p.category),
            space,
            METHOD_LABELS.get(p.payment_method, p.payment_method),
            _cop(p.amount),
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = _THIN_BORDER
            cell.font = _font(10)
            cell.alignment = _align("right") if c == 7 else _align("left")
            if c == 7:
                cell.number_format = '#,##0'
            if i % 2 == 0:
                cell.fill = _fill("F5F5F5")
        total += p.amount

    total_row = len(filtrado) + 2
    bg = "C8E6C9" if tipo == "income" else "FFCDD2"
    txt_color = "1B5E20" if tipo == "income" else "B71C1C"
    _style_total_row(ws, total_row, len(headers), bg=bg)
    ws.merge_cells(f"A{total_row}:F{total_row}")
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = _font(10, bold=True)
    ws[f"A{total_row}"].alignment = _align("right")
    ws[f"A{total_row}"].fill = _fill(bg)
    ws[f"A{total_row}"].border = _THIN_BORDER
    cell = ws.cell(row=total_row, column=7, value=_cop(total))
    cell.number_format = '#,##0'
    cell.font = _font(11, bold=True, color=txt_color)
    cell.alignment = _align("right")
    cell.fill = _fill(bg)
    cell.border = _THIN_BORDER

    _autofit(ws)


def _build_kpis(ws, payments: list[Payment], space_names: dict):
    """Hoja 4: KPIs del período — single pass over payments."""
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:C1")
    ws["A1"] = "KPIs DEL PERÍODO"
    ws["A1"].font = _font(12, bold=True, color="1F4E79")
    ws["A1"].alignment = _align("center")
    ws["A1"].fill = _fill("E3F2FD")

    for c, h in enumerate(["Indicador", "Valor", "Nota"], 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, 3)

    # Single pass aggregation
    total_inc = Decimal(0)
    total_exp = Decimal(0)
    inc_count = exp_count = 0
    inc_by_cat: dict[str, Decimal] = {}
    exp_by_cat: dict[str, Decimal] = {}
    space_inc: dict[str, Decimal] = {}
    space_exp: dict[str, Decimal] = {}

    for p in payments:
        space = space_names.get(p.space_id, "General") if p.space_id else "General"
        label = CATEGORY_LABELS.get(p.category, p.category)
        if p.type == "income":
            total_inc += p.amount
            inc_count += 1
            inc_by_cat[label] = inc_by_cat.get(label, Decimal(0)) + p.amount
            space_inc[space] = space_inc.get(space, Decimal(0)) + p.amount
        else:
            total_exp += p.amount
            exp_count += 1
            exp_by_cat[label] = exp_by_cat.get(label, Decimal(0)) + p.amount
            space_exp[space] = space_exp.get(space, Decimal(0)) + p.amount

    neto = total_inc - total_exp
    margen = float(neto / total_inc * 100) if total_inc else 0.0
    ticket_prom = _cop(total_inc / inc_count) if inc_count else 0

    kpis: list[tuple] = [
        ("Total ingresos del período", _cop(total_inc), "COP"),
        ("Total egresos del período", _cop(total_exp), "COP"),
        ("Balance neto", _cop(neto), "COP — positivo = utilidad"),
        ("Margen operacional", f"{margen:.1f}%", "Neto / Ingresos × 100"),
        ("Número de ingresos registrados", inc_count, "transacciones"),
        ("Número de egresos registrados", exp_count, "transacciones"),
        ("Ticket promedio por ingreso", ticket_prom, "COP por transacción"),
    ]

    for cat, val in sorted(inc_by_cat.items(), key=lambda x: x[1], reverse=True):
        pct = float(val / total_inc * 100) if total_inc else 0.0
        kpis.append((f"  Ingreso — {cat}", _cop(val), f"{pct:.1f}% del total ingresos"))

    for cat, val in sorted(exp_by_cat.items(), key=lambda x: x[1], reverse=True):
        pct = float(val / total_exp * 100) if total_exp else 0.0
        kpis.append((f"  Egreso — {cat}", _cop(val), f"{pct:.1f}% del total egresos"))

    for sp in sorted(set(list(space_inc) + list(space_exp))):
        inc_sp = space_inc.get(sp, Decimal(0))
        exp_sp = space_exp.get(sp, Decimal(0))
        kpis += [
            (f"  {sp} — Ingresos", _cop(inc_sp), "COP"),
            (f"  {sp} — Egresos", _cop(exp_sp), "COP"),
            (f"  {sp} — Neto", _cop(inc_sp - exp_sp), "COP"),
        ]

    for i, (nombre, valor, nota) in enumerate(kpis, 3):
        bg = "BBDEFB" if "Balance neto" in nombre else ("F5F5F5" if i % 2 == 0 else "FFFFFF")
        ws.cell(row=i, column=1, value=nombre).font = _font(10)
        val_cell = ws.cell(row=i, column=2, value=valor)
        val_cell.font = _font(10, bold=True)
        val_cell.alignment = _align("right")
        if isinstance(valor, int):
            val_cell.number_format = '#,##0'
        ws.cell(row=i, column=3, value=nota).font = _font(9, color="666666")
        for c in range(1, 4):
            ws.cell(row=i, column=c).fill = _fill(bg)
            ws.cell(row=i, column=c).border = _THIN_BORDER

    _autofit(ws)


def _build_por_espacio(ws, payments: list[Payment], space_names: dict):
    """Hoja 5: Desglose por espacio."""
    ws.freeze_panes = "A2"

    headers = ["Espacio", "Ingresos (COP)", "Egresos (COP)", "Neto (COP)", "% Ingresos totales"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    space_data: dict[str, dict] = {}
    for p in payments:
        space = space_names.get(p.space_id, "General") if p.space_id else "General"
        if space not in space_data:
            space_data[space] = {"income": Decimal(0), "expense": Decimal(0)}
        space_data[space][p.type] += p.amount

    total_inc_all = sum(d["income"] for d in space_data.values())

    for i, (space, data) in enumerate(sorted(space_data.items()), 2):
        inc = data["income"]
        exp = data["expense"]
        neto = inc - exp
        pct = float(inc / total_inc_all * 100) if total_inc_all else 0.0
        neto_color = "1B5E20" if neto >= 0 else "B71C1C"
        bg = "F5F5F5" if i % 2 == 0 else "FFFFFF"
        for c, v in enumerate([space, _cop(inc), _cop(exp), _cop(neto), f"{pct:.1f}%"], 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = _THIN_BORDER
            cell.fill = _fill(bg)
            if c == 1:
                cell.font = _font(10, bold=True)
                cell.alignment = _align("left")
            elif c == 4:
                cell.font = _font(10, bold=True, color=neto_color)
                cell.alignment = _align("right")
                cell.number_format = '#,##0'
            else:
                cell.font = _font(10)
                cell.alignment = _align("right")
                if c in (2, 3):
                    cell.number_format = '#,##0'

    total_row = len(space_data) + 2
    total_exp_all = sum(d["expense"] for d in space_data.values())
    grand_neto = total_inc_all - total_exp_all
    _style_total_row(ws, total_row, len(headers), bg="E8EAF6")
    ws.cell(row=total_row, column=1, value="TOTAL").font = _font(10, bold=True)
    for c, v in enumerate([_cop(total_inc_all), _cop(total_exp_all), _cop(grand_neto), "100%"], 2):
        cell = ws.cell(row=total_row, column=c, value=v)
        cell.font = _font(10, bold=True)
        cell.alignment = _align("right")
        cell.fill = _fill("E8EAF6")
        cell.border = _THIN_BORDER
        if c in (2, 3, 4) and isinstance(v, int):
            cell.number_format = '#,##0'

    _autofit(ws)


@router.get("/contabilidad")
async def export_contabilidad(
    start: Optional[str] = Query(None, description="Fecha inicio YYYY-MM-DD"),
    end: Optional[str] = Query(None, description="Fecha fin YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """
    Genera y descarga un archivo Excel con el reporte contable del período.
    Hojas: Resumen Ejecutivo | Ingresos Detalle | Egresos Detalle | KPIs | Por Espacio
    """
    if not current_user.tenant_id:
        raise HTTPException(403, "Sin tenant")

    q = select(Payment).where(Payment.tenant_id == current_user.tenant_id)
    if start:
        q = q.where(Payment.payment_date >= date.fromisoformat(start))
    if end:
        q = q.where(Payment.payment_date <= date.fromisoformat(end))
    result = await db.execute(q.order_by(Payment.payment_date.asc()))
    payments_list = result.scalars().all()

    spaces_result = await db.execute(select(Space).where(Space.tenant_id == current_user.tenant_id))
    space_names: dict[uuid.UUID, str] = {s.id: s.name for s in spaces_result.scalars().all()}

    client_ids = {p.client_id for p in payments_list if p.client_id}
    client_names: dict[uuid.UUID, str] = {}
    if client_ids:
        clients_result = await db.execute(select(Client).where(Client.id.in_(list(client_ids))))
        client_names = {c.id: c.full_name for c in clients_result.scalars().all()}

    tenant = await db.get(Tenant, current_user.tenant_id)
    tenant_name = tenant.name if tenant else "Estudio"

    wb = openpyxl.Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Ejecutivo"
    ws_ingresos = wb.create_sheet("Ingresos Detalle")
    ws_egresos = wb.create_sheet("Egresos Detalle")
    ws_kpis = wb.create_sheet("KPIs del Período")
    ws_espacios = wb.create_sheet("Por Espacio")

    _build_resumen(ws_resumen, payments_list, tenant_name, start or "—", end or "—", space_names)
    _build_detalle(ws_ingresos, payments_list, "income", client_names, space_names)
    _build_detalle(ws_egresos, payments_list, "expense", client_names, space_names)
    _build_kpis(ws_kpis, payments_list, space_names)
    _build_por_espacio(ws_espacios, payments_list, space_names)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    start_label = (start or "inicio").replace("-", "")
    end_label = (end or "fin").replace("-", "")
    filename = f"Contabilidad_{tenant_name.replace(' ', '_')}_{start_label}_{end_label}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
